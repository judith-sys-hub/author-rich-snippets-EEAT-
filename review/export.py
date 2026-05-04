import json
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

DATA_DIR = Path(__file__).parent.parent / "data" / "authors"
XLSX_PATH = Path(__file__).parent / "author_review.xlsx"


def build_row(author_data: dict) -> dict:
    profile = author_data["profile"]
    beats = (author_data.get("expertise") or {}).get("beats") or []
    return {
        "slug": author_data["meta"]["slug"],
        "name": profile.get("name", ""),
        "status": "pending",
        "bio": profile.get("bio_generated") or "",
        "themen": "; ".join(beats),
    }


def write_xlsx(rows: list[dict], xlsx_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Autorenreview"

    last_row = max(len(rows) + 2, 3)  # row 1 summary + row 2 header + data rows; min 3 so ranges are valid

    # Row 1: summary with live COUNTIF formulas
    summary_fill = PatternFill("solid", fgColor="E3F2FD")
    summary_font = Font(bold=True)
    ws["A1"] = "Fortschritt"
    ws["B1"] = f"=COUNTA(A3:A{last_row})"
    ws["C1"] = f'=COUNTIF(C3:C{last_row},"approved")'
    ws["D1"] = f'=COUNTIF(C3:C{last_row},"pending")'
    ws["E1"] = f'=COUNTIF(C3:C{last_row},"flagged")'
    for col in "ABCDE":
        ws[f"{col}1"].fill = summary_fill
        ws[f"{col}1"].font = summary_font

    # Row 2: headers (frozen)
    header_fill = PatternFill("solid", fgColor="E8EAF6")
    header_font = Font(bold=True)
    for i, col_name in enumerate(["slug", "name", "status", "bio", "themen"], start=1):
        cell = ws.cell(row=2, column=i, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A3"

    # Data rows
    grey_font = Font(color="BBBBBB")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for r_idx, row in enumerate(rows, start=3):
        ws.cell(row=r_idx, column=1, value=row["slug"]).font = grey_font
        ws.cell(row=r_idx, column=2, value=row["name"])
        ws.cell(row=r_idx, column=3, value=row["status"])
        bio_cell = ws.cell(row=r_idx, column=4, value=row["bio"])
        bio_cell.alignment = wrap_align
        ws.cell(row=r_idx, column=5, value=row["themen"])
        ws.row_dimensions[r_idx].height = 60

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 35

    # Conditional formatting on column C
    cf_range = f"C3:C{last_row}"
    ws.conditional_formatting.add(
        cf_range,
        CellIsRule(operator="equal", formula=['"approved"'], fill=PatternFill("solid", fgColor="C8E6C9")),
    )
    ws.conditional_formatting.add(
        cf_range,
        CellIsRule(operator="equal", formula=['"flagged"'], fill=PatternFill("solid", fgColor="FFCDD2")),
    )
    ws.conditional_formatting.add(
        cf_range,
        CellIsRule(operator="equal", formula=['"pending"'], fill=PatternFill("solid", fgColor="FFF9C4")),
    )

    # Data validation dropdown on column C
    dv = DataValidation(type="list", formula1='"pending,approved,flagged"', allow_blank=False)
    dv.sqref = f"C3:C{last_row}"
    ws.add_data_validation(dv)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def export_authors(data_dir: Path = DATA_DIR, xlsx_path: Path = XLSX_PATH) -> int:
    from scraper.state import get_authors_by_stage

    candidates = list(get_authors_by_stage("rendered"))
    rows = []
    for candidate in candidates:
        slug = candidate["slug"]
        json_path = data_dir / f"{slug}.json"
        if not json_path.exists():
            print(f"[export] WARNUNG: {json_path.name} nicht gefunden – uebersprungen")
            continue
        author_data = json.loads(json_path.read_text(encoding="utf-8"))
        try:
            rows.append(build_row(author_data))
        except Exception as exc:
            raise RuntimeError(f"Fehler bei {slug}: {exc}") from exc

    write_xlsx(rows, xlsx_path)
    return len(rows)


if __name__ == "__main__":
    count = export_authors()
    print(f"[export] {count} Autoren exportiert -> {XLSX_PATH}")
