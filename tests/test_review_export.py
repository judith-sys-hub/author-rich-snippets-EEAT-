import json
from pathlib import Path
from unittest.mock import patch

import openpyxl

from review.export import build_row, export_authors, write_xlsx


def _make_author(slug="test-autor", beats=None, bio="Ein Redakteur."):
    if beats is None:
        beats = ["Sport", "Fussball"]
    return {
        "meta": {"slug": slug},
        "profile": {"name": "Test Autor", "bio_generated": bio},
        "expertise": {"beats": beats, "derived_from_articles": True},
        "articles": [],
        "enrichment": {"awards": [], "social_links": {}, "wikipedia_url": None},
    }


def _read_data_rows(xlsx_path: Path) -> list[dict]:
    """Read data rows (row 3+) as list of dicts, skipping summary and header."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []
    for row_vals in ws.iter_rows(min_row=3, values_only=True):
        padded = list(row_vals) + [None] * 5
        slug, name, status, bio, themen = padded[:5]
        if slug:
            rows.append({"slug": slug, "name": name, "status": status, "bio": bio, "themen": themen})
    return rows


# --- build_row unit tests ---

def test_build_row_slug():
    assert build_row(_make_author(slug="maria-muster"))["slug"] == "maria-muster"


def test_build_row_name():
    assert build_row(_make_author())["name"] == "Test Autor"


def test_build_row_status_is_pending():
    assert build_row(_make_author())["status"] == "pending"


def test_build_row_bio():
    assert build_row(_make_author(bio="Eine Journalistin."))["bio"] == "Eine Journalistin."


def test_build_row_themen_semicolon_joined():
    assert build_row(_make_author(beats=["Sport", "Fussball", "Tennis"]))["themen"] == "Sport; Fussball; Tennis"


def test_build_row_no_beats_empty_themen():
    assert build_row(_make_author(beats=[]))["themen"] == ""


def test_build_row_missing_expertise():
    author = _make_author()
    del author["expertise"]
    assert build_row(author)["themen"] == ""


def test_build_row_null_beats():
    author = _make_author()
    author["expertise"]["beats"] = None
    assert build_row(author)["themen"] == ""


def test_build_row_null_bio():
    author = _make_author()
    author["profile"]["bio_generated"] = None
    assert build_row(author)["bio"] == ""


# --- write_xlsx unit tests ---

def test_write_xlsx_creates_file(tmp_path):
    xlsx_path = tmp_path / "test.xlsx"
    write_xlsx([{"slug": "a", "name": "A", "status": "pending", "bio": "Bio.", "themen": "Sport"}], xlsx_path)
    assert xlsx_path.exists()


def test_write_xlsx_header_row(tmp_path):
    xlsx_path = tmp_path / "test.xlsx"
    write_xlsx([], xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "slug"
    assert ws.cell(row=2, column=3).value == "status"
    assert ws.cell(row=2, column=5).value == "themen"


def test_write_xlsx_summary_row_has_countif_formula(tmp_path):
    xlsx_path = tmp_path / "test.xlsx"
    write_xlsx([{"slug": "a", "name": "A", "status": "pending", "bio": "Bio.", "themen": "Sport"}], xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)  # no data_only — need formula text
    ws = wb.active
    assert str(ws["C1"].value).startswith("=COUNTIF")


def test_write_xlsx_data_row_values(tmp_path):
    xlsx_path = tmp_path / "test.xlsx"
    rows = [{"slug": "maria-muster", "name": "Maria Muster", "status": "pending", "bio": "Eine Bio.", "themen": "Sport; Fussball"}]
    write_xlsx(rows, xlsx_path)
    data = _read_data_rows(xlsx_path)
    assert data[0]["slug"] == "maria-muster"
    assert data[0]["status"] == "pending"
    assert data[0]["bio"] == "Eine Bio."
    assert data[0]["themen"] == "Sport; Fussball"


def test_write_xlsx_conditional_formatting_present(tmp_path):
    xlsx_path = tmp_path / "test.xlsx"
    write_xlsx([{"slug": "a", "name": "A", "status": "pending", "bio": "Bio.", "themen": ""}], xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert len(list(ws.conditional_formatting)) > 0


def test_write_xlsx_data_validation_present(tmp_path):
    xlsx_path = tmp_path / "test.xlsx"
    write_xlsx([{"slug": "a", "name": "A", "status": "pending", "bio": "Bio.", "themen": ""}], xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert len(ws.data_validations.dataValidation) > 0


# --- export_authors integration tests ---

def test_export_returns_count(tmp_path):
    data_dir = tmp_path / "authors"
    data_dir.mkdir()
    xlsx_path = tmp_path / "author_review.xlsx"
    author = _make_author(slug="test-autor")
    (data_dir / "test-autor.json").write_text(json.dumps(author, ensure_ascii=False), encoding="utf-8")
    with patch("scraper.state.get_authors_by_stage", return_value=[{"slug": "test-autor"}]):
        count = export_authors(data_dir=data_dir, xlsx_path=xlsx_path)
    assert count == 1


def test_export_writes_xlsx_file(tmp_path):
    data_dir = tmp_path / "authors"
    data_dir.mkdir()
    xlsx_path = tmp_path / "author_review.xlsx"
    author = _make_author(slug="test-autor")
    (data_dir / "test-autor.json").write_text(json.dumps(author, ensure_ascii=False), encoding="utf-8")
    with patch("scraper.state.get_authors_by_stage", return_value=[{"slug": "test-autor"}]):
        export_authors(data_dir=data_dir, xlsx_path=xlsx_path)
    assert xlsx_path.exists()


def test_export_row_values(tmp_path):
    data_dir = tmp_path / "authors"
    data_dir.mkdir()
    xlsx_path = tmp_path / "author_review.xlsx"
    author = _make_author(slug="test-autor", beats=["Sport", "Fussball"], bio="Ein Sportredakteur.")
    (data_dir / "test-autor.json").write_text(json.dumps(author, ensure_ascii=False), encoding="utf-8")
    with patch("scraper.state.get_authors_by_stage", return_value=[{"slug": "test-autor"}]):
        export_authors(data_dir=data_dir, xlsx_path=xlsx_path)
    rows = _read_data_rows(xlsx_path)
    assert rows[0]["slug"] == "test-autor"
    assert rows[0]["status"] == "pending"
    assert rows[0]["bio"] == "Ein Sportredakteur."
    assert rows[0]["themen"] == "Sport; Fussball"


def test_export_overwrites_existing_file(tmp_path):
    data_dir = tmp_path / "authors"
    data_dir.mkdir()
    xlsx_path = tmp_path / "author_review.xlsx"
    xlsx_path.write_bytes(b"stale")
    author = _make_author(slug="test-autor")
    (data_dir / "test-autor.json").write_text(json.dumps(author, ensure_ascii=False), encoding="utf-8")
    with patch("scraper.state.get_authors_by_stage", return_value=[{"slug": "test-autor"}]):
        export_authors(data_dir=data_dir, xlsx_path=xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    assert wb.active.cell(row=2, column=1).value == "slug"


def test_export_skips_missing_json_file(tmp_path):
    data_dir = tmp_path / "authors"
    data_dir.mkdir()
    xlsx_path = tmp_path / "author_review.xlsx"
    with patch("scraper.state.get_authors_by_stage", return_value=[{"slug": "ghost-autor"}]):
        count = export_authors(data_dir=data_dir, xlsx_path=xlsx_path)
    assert count == 0
